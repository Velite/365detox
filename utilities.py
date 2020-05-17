from dateutil import parser, tz
import msgpack
from openpyxl import load_workbook


class DateTime:
	@classmethod
	def parse2date(cls, value):
		return parser.parse(value, dayfirst = True) if type(value) is str else value

	@classmethod
	def date2utc(cls, value):
		parseddate = cls.parse2date(value)
		return parseddate.replace(tzinfo = tz.tzutc()) if parseddate.tzinfo is None else parseddate.astimezone(
			tz.tzutc())

	@classmethod
	def date2local(cls, value):
		parseddate = cls.parse2date(value)
		return parseddate.replace(tzinfo = tz.tzlocal()) if parseddate.tzinfo is None else parseddate.astimezone(
			tz.tzlocal())

	@classmethod
	def utc2local(cls, value):
		return cls.date2utc(value).astimezone(tz.tzlocal())

	@classmethod
	def local2utc(cls, value):
		return cls.date2local(value).astimezone(tz.tzutc())

	@classmethod
	def date2string(cls, value):
		return cls.parse2date(value).strftime("%d.%m.%Y")

	@classmethod
	def datetime2string(cls, value):
		return cls.parse2date(value).strftime("%d.%m.%Y %H:%M:%S")


class MessagePacket:
	@classmethod
	def generate(cls, value):
		return [b for b in msgpack.packb(value)]

	@classmethod
	def encode(cls, value):
		return msgpack.unpackb(value, encoding = "utf-8")


class Excel:
	@classmethod
	def get_headers(cls, path):
		wb = load_workbook(filename = path, read_only = True)
		sh = wb.worksheets[0]
		return [dict(Name = sh.cell(row = 1, column = r + 1).value, Index = r + 1) for r in range(sh.max_column)]

	@classmethod
	def get_data(cls, path, fields):
		wb = load_workbook(filename = path, read_only = True)
		sh = wb.worksheets[0]
		return [dict(Name = sh.cell(row = c + 1, column = int(fields.get("Name"))).value,
		             Barcode = sh.cell(row = c + 1, column = int(fields.get("BarCode"))).value,
		             StartSale = sh.cell(row = c + 1, column = int(fields.get("StartSaleDate"))).value,
		             Position = 1000.0) for c in
		        range(1, sh.max_row)]
